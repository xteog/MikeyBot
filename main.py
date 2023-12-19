from datetime import datetime
from datetime import timedelta

import openpyxl
import utils
from typing import Any
import discord
from discord.ext import commands
import slashCommands
import moderation.views
import moderation.moderation

import config
import logging
import sys


class MyBot(commands.Bot):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        if sys.argv[1] == "run":
            config.RUN()
        elif sys.argv[1] == "test":
            config.TEST()
        else:
            raise ValueError
        self.server = config.serverId
        self.dmsChannel = None
        self.errorChannel = None
        self.reportChannel = None
        self.voiceClient = None
        self.lastAnnouncement = {}

    async def on_ready(self):
        await self.add_cog(
            slashCommands.setup(self), guild=discord.Object(id=self.server)
        )
        await self.tree.sync(guild=discord.Object(id=self.server))
        self.errorChannel = self.get_channel(config.errorChannelId)
        self.reportChannel = self.get_channel(config.reportChannelId)

        reports = await moderation.moderation.getActive(self)

        for r in reports:
            view = moderation.views.ReportView(bot=self, data=r)
            self.add_view(view)
            
        print("Mikey is up")

        self.ready = True

    async def on_message(self, message: discord.Message):
        if message.author == self.user:
            return

        await self.devCommands(message)

        if message.channel.id == self.reportChannel.id:
            await self.deleteMessage(self.reportChannel.id, message.id)

        for league in config.openTime.keys():
            if datetime.now() > config.openTime[league]:
                if not (league in self.lastAnnouncement.keys()):
                    data = utils.read("data/lastAnnouncement.json")
                    if data == None or not (league in data.keys()):
                        self.lastAnnouncement[league] = datetime.strptime(
                            "2001-11-09 8:09", config.timeFormat
                        )

                        utils.write(
                            "data/lastAnnouncement.json",
                            {
                                key: datetime.strftime(
                                    self.lastAnnouncement[key], config.timeFormat
                                )
                                for key in self.lastAnnouncement.keys()
                            },
                        )
                    else:
                        self.lastAnnouncement[league] = datetime.strptime(
                            data[league], config.timeFormat
                        )

                if self.lastAnnouncement[league] < config.openTime[league]:
                    msg = f"Reports window is now open until <t:{int(config.closeTime[league].timestamp())}:f>"
                    await self.sendMessage(msg, config.leaguesChannelIds[league])

                    self.lastAnnouncement[league] = datetime.now()
                    utils.write(
                        "data/lastAnnouncement.json",
                        {
                            key: datetime.strftime(
                                self.lastAnnouncement[key], config.timeFormat
                            )
                            for key in self.lastAnnouncement.keys()
                        },
                    )

        if isinstance(message.channel, discord.DMChannel):
            logging.info(
                f"DM by {message.author.name} ({message.channel.id}): {message.content}"
            )

    async def devCommands(self, message: discord.Message):
        if (
            message.content.startswith("$send")
            and message.author.id == 493028834640396289
        ):
            str = message.content.split("-")
            try:
                channel = await self.fetch_channel(str[1])
                await channel.send(str[2])
            except:
                await self.errorChannel.send("Error during sending message")

        if (
            message.content.startswith("$delete")
            and message.author.id == 493028834640396289
        ):
            str = message.content.split("-")
            try:
                await self.deleteMessage(str[1], str[2])
            except:
                await self.errorChannel.send("Error during replying message")

        if (
            message.content.startswith("$reply")
            and message.author.id == 493028834640396289
        ):
            str = message.content.split("-")
            try:
                channel = await self.fetch_channel(str[1])
                msg = await channel.fetch_message(str[2])
                await msg.reply(str[3])
            except:
                await self.errorChannel.send("Error during replying message")

        if (
            message.content.startswith("$logs")
            and message.author.id == 493028834640396289
        ):
            file = discord.File("./data/logging.log")
            await message.channel.send(file=file)

        if message.content.startswith("$history") and (
            message.author.id == 493028834640396289
            or message.author.id == 1181521363127767130
        ):
            file = discord.File("./data/history.json")
            await self.reportChannel.send(file=file)

    async def sendReport(self, data: moderation.moderation.ReportData):
        view = moderation.views.ReportView(self, data)
        message = await self.reportChannel.send(embed=view.embed, view=view)
        await self.reportChannel.create_thread(
            name=f"Report {data.offender.name} ({data.id})",
            message=message,
            auto_archive_duration=1440,
        )

    async def sendReminder(self, data: moderation.moderation.ReportData) -> None:
        embed = moderation.views.ReportEmbed(data, permission=False)
        await data.offender.send(embed=embed)

    async def deleteMessage(self, channelId: int, messageId: int) -> None:
        channel = await self.fetch_channel(channelId)
        msg = await channel.fetch_message(messageId)
        await msg.delete()

    async def sendMessage(self, msg: str, channelId: int) -> None:
        channel = self.get_channel(channelId)
        await channel.send(msg)

    async def archiveThread(self, id: str) -> None:
        threads = self.reportChannel.threads
        for t in threads:
            if t.name.find(id) != -1:
                await t.edit(archived=True)

    async def writeWorkbook(self):
        history = utils.read(config.historyPath)
        workbook = openpyxl.load_workbook(filename=config.penaltyLogPath)
        for member in history.keys():
            for id in history[member]["violations"].keys():
                data = (await moderation.moderation.getReports(self, id=id))[0]
                
                sheet = workbook.active

                i = 1
                while True:
                    if sheet.cell(row=i, column=1).value == None:
                        sheet.cell(row=i, column=1).value = data.id
                        sheet.cell(row=i, column=2).value = data.offender.name
                        sheet.cell(row=i, column=3).value = data.penalty
                        sheet.cell(row=i, column=4).value = data.severity
                        sheet.cell(row=i, column=5).value = data.league
                        sheet.cell(row=i, column=6).value = data.round
                        if len(str(data.rule)) > 0:
                            sheet.cell(row=i, column=7).value = str(data.rule)
                            sheet.cell(row=i, column=8).value = data.proof
                            sheet.cell(row=i, column=9).value = data.notes
                        else:
                            sheet.cell(row=i, column=9).value = str(data.rule)
                            sheet.cell(row=i, column=8).value = data.proof
                            sheet.cell(row=i, column=7).value = data.notes
                        sheet.cell(row=i, column=10).value = data.creator.name
                        sheet.cell(row=i, column=11).value = data.desc
                        sheet.cell(row=i, column=12).value = data.timestamp
                        break
                    i += 1

        workbook.save(filename=config.penaltyLogPath)

    async def on_error(self, event_method: str, /, *args: Any, **kwargs: Any) -> None:
        logging.error(f"Error: {event_method}")


def runBot():
    intents = discord.Intents.default()
    intents.message_content = True
    bot = MyBot(intents=intents, command_prefix=".")
    bot.run(config.Token, reconnect=True)


async def reconnect(bot):
    await bot.close()
    runBot()


if __name__ == "__main__":
    logging.basicConfig(
        filename="data/logging.log",
        format="%(asctime)s [%(levelname)s]:%(name)s:%(message)s",
        level=logging.INFO,
    )

    runBot()
