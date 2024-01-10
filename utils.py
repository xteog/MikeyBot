import json
import random
import discord
import socket
import codecs
import openpyxl


def lev_dist(s: str, t: str) -> int:
    """
    Calculates the lev distance between two strings.

    Parameters
    -----------
    - s : `str`\n
        First string to confront.
    - t : `str`\n
        Second string to confront.

    Returns
    ----------
    A `int` value representing the lev distance between the two strings.
    """
    v0 = [i for i in range(len(t) + 1)]
    v1 = [0] * (len(t) + 1)

    for i in range(len(s)):
        v1[0] = i + 1
        for j in range(len(t)):
            deletionCost = v0[j + 1] + 1
            insertionCost = v1[j] + 1
            if s[i] == t[j]:
                substitutionCost = v0[j]
            else:
                substitutionCost = v0[j] + 1

            v1[j + 1] = min(deletionCost, insertionCost, substitutionCost)

        temp = v0
        v0 = v1
        v1 = temp
    return v0[-1]


def loading(i, len):
    j = 0
    str = "|"
    perc = round(i / len * 20, 1)
    while j < 20:
        if perc > j:
            str += "█"
        else:
            str += "   "
        j += 1
    str += f"| {round(i/len*100, 1)}%"
    return str


def updateWorkbook(path: str, data) -> None:
    workbook = openpyxl.load_workbook(filename=path)
    sheet = workbook.active

    i = 1
    while True:
        if sheet.cell(row=i, column=1).value == None:
            sheet.cell(row=i, column=1).value = data.id
            sheet.cell(row=i, column=2).value = data.offender.name
            sheet.cell(row=i, column=3).value = data.penalty
            sheet.cell(row=i, column=4).value = data.severity
            sheet.cell(row=i, column=5).value = data.league
            sheet.cell(row=i, column=6).value = data.round
            sheet.cell(row=i, column=7).value = str(data.rule)
            sheet.cell(row=i, column=8).value = data.proof
            sheet.cell(row=i, column=9).value = data.notes
            sheet.cell(row=i, column=10).value = data.creator.name
            sheet.cell(row=i, column=11).value = data.desc
            sheet.cell(row=i, column=12).value = data.timestamp
            break
        i += 1

    workbook.save(filename=path)

def createWorkbook(path: str, data:list[list]):
    workbook = openpyxl.load_workbook(filename=path)
    sheet = workbook.active

    for i in range(len(data)):
        for j in range(len(data[0])):
            sheet.cell(row=i + 1, column=j + 1).value = data[i][j]

    workbook.save(filename=path)


def write(path: str, data: dict) -> None:
    data = json.dumps(data, indent=2)
    with open(path, "w+") as f:
        f.write(data)


def read(path: str) -> dict:
    try:
        with open(path, "r") as f:
            file = f.read()
        return json.loads(file)
    except:
        return None


def randomString(len: int) -> str:
    alphabet = "0123456789"

    return random.choices(alphabet[1:])[0] + "".join(
        random.choices(alphabet, k=len - 1)
    )


def isLink(str: str) -> dict[bool, str]:
    """
    Checks if the string provided is a YouTube or Discord link.

    Parameters
    -----------
    - str : `str`\n
        The string to check.

    Returns
    ----------
    - `bool` value representing the outcome of the check.

    - `str` containing the reason why the check failed.
    """
    cond = len(str) > 5 and str.find(" ") == -1

    yt = cond and (
        str.startswith("https://youtube.com")
        or str.startswith("https://www.youtube.com")
        or str.startswith("youtube.com")
        or str.startswith("https://youtu.be")
        or str.startswith("https://www.youtu.be")
        or str.startswith("youtu.be")
    )

    discord = cond and (
        str.startswith("https://discord.com")
        or str.startswith("discord.com")
        or str.startswith("https://cdn.discordapp.com")
        or str.startswith("cdn.discordapp.com")
    )

    if yt:
        if linkHasTimestamp(str):
            return True, None
        else:
            return False, "The YouTube link provided doesn't have a timestamp"
    elif discord:
        return True, None

    return False, "The link provided ain't a YouTube or Discord link"


def linkHasTimestamp(str: str) -> bool:
    start = str.find("t=")

    if start == -1:
        return False

    if len(str[start:]) <= 3:
        return False

    return str[start + 3 :].isdigit() or str[len(str) - 1] == "s"


def hasPermissions(
    user: discord.Member, role: int = None, roles: list[int] = None
) -> bool:
    """
    Checks if the provided `Member` has a set of roles.

    Only one of the parameters `role` or `roles` should be given.

    Parameters
    -----------
    - user : `discord.Member`\n
        User to check if has permissions.
    - role : `Optional[int]`
        The role code that the user should have.
    - roles : `Optional[list[int]]`
        The roles codes that the user should have.

    Returns
    ----------
    `bool`
        the outcome of the check.
    """
    if role != None:
        roles = [role]

    for i in roles:
        for j in user.roles:
            if j.id == i:
                return True

    return False


def getLobbyInfo(data: str) -> dict:
    """
    This is a reST style.

    :param str: this is a first param
    :returns: this is a description of what is returned
    """
    start, end = 0, 0
    for i in range(len(data)):
        if data[i] == "#":
            start = i + 1
        if data[i] == "|":
            end = i
            break
    lobbyName = data[start:end]

    start, end, cont = 0, 0, 0
    for i in range(len(data)):
        if data[i] == "|":
            if start == 0:
                cont += 1
            else:
                end = i
                break

        if data[i] == "|" and cont == 3:
            start = i + 1
    max_players = data[start:end]

    start, end, cont = 0, 0, 0
    for i in range(len(data)):
        if data[i] == "|":
            if start == 0:
                cont += 1
            else:
                end = i
                break

        if data[i] == "|" and cont == 9:
            start = i + 1
    players = data[start:end]

    start, end, cont = 0, 0, 0
    for i in range(len(data)):
        if data[i] == "|":
            if start == 0:
                cont += 1
            else:
                end = i
                break

        if data[i] == "|" and cont == 7:
            start = i + 1
    status = data[start:end]

    start, end, cont = 0, 0, 0
    for i in range(len(data)):
        if data[i] == "|":
            if start == 0:
                cont += 1
            else:
                end = i
                break

        if data[i] == "|" and cont == 5:
            start = i + 1
    curr_laps = data[start:end]

    start, end, cont = 0, 0, 0
    for i in range(len(data)):
        if data[i] == "|":
            if start == 0:
                cont += 1
            else:
                end = i
                break

        if data[i] == "|" and cont == 6:
            start = i + 1
    max_laps = data[start:end]

    start, end, cont = 0, 0, 0
    for i in range(len(data)):
        if data[i] == "|":
            if start == 0:
                cont += 1
            else:
                end = i
                break

        if data[i] == "|" and cont == 10:
            start = i + 1
    password = data[start:end]

    if lobbyName != "*0":
        return {
            "name": lobbyName,
            "curr_players": players,
            "max_players": max_players,
            "status": status,
            "curr_laps": curr_laps,
            "max_laps": max_laps,
            "private": password != "",
        }
    else:
        return None


def getLobbiesList() -> dict:
    lobbies = []
    IP = "46.101.147.176"
    PORT = 6510
    payload = f"39300a00"

    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    sock.bind(("127.0.0.1", PORT))

    data = codecs.decode(payload, "hex_codec")
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    sock.sendto(data, (IP, PORT))

    run = True
    sock.settimeout(2)
    while run:
        try:
            data, addr = sock.recvfrom(1024)
            data = data.decode()

            lobby = getLobbyInfo(data)
            if lobby != None:
                lobbies.append(lobby)
        except:
            run = False

    sock.close()

    return lobbies

def formatBlockQuote(str) -> str:
    str = "> " + str

    for i in range(len(str)):
        if str[i] == "\n":
            str = str[:i] + "> " + str[i:]

    return str    